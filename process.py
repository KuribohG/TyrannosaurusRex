from openpyxl import load_workbook
from functools import reduce
from collections import defaultdict
import argparse
import os

parser = argparse.ArgumentParser()
parser.add_argument('--separate', action='store_true')

data_wb = load_workbook('b.xlsx', data_only=True)
ids = data_wb.sheetnames

def bin_by_date_and_id(data):
	ret = defaultdict(lambda: defaultdict(list))
	for row in data:
		machine_id = row[1].value
		date = row[2].value
		ret[date][machine_id].append(row)
	return ret

def get_row(page, offset):
	return page * 24 + offset + 5

def sample_num(x):
	if x <= 15:
		return 2
	elif x <= 25:
		return 3
	elif x <= 90:
		return 5
	elif x <= 150:
		return 8
	elif x <= 280:
		return 13
	elif x <= 500:
		return 20
	elif x <= 1200:
		return 32
	else:
		return 50

def uids_to_str(uids):
	d = defaultdict(list)
	for uid in uids:
		d[uid[:3]].append(int(uid[3:]))
	ret = ''
	first_k = True
	for k, l in d.items():
		x = sorted(l)
		cur = 0
		r = 0
		if first_k:
			ret += k
		else:
			ret += '；' + k
		first_k = False
		first_n = True
		while cur < len(x):
			r = cur
			while r + 1 < len(x) and x[r + 1] <= x[r] + 1:
				r += 1
			low = x[cur]
			high = x[r]
			to_append = '' if first_n else '、'
			if low == high:
				to_append += '{}'.format(low)
			else:
				to_append += '{}~{}'.format(low, high)
			first_n = False
			ret += to_append
			cur = r + 1
	return ret

def date_to_chinese(s):
	l = s.split('.')
	return l[0] + '年' + l[1] + '月' + l[2] + '日'

def modify(date, v, sheet_no, baoshen, shigong, jianyan, yinbi):
	sheet_no_str = '{:03d}'.format(sheet_no)
	baoshen['L5'] = sheet_no_str
	jianyan['BF4'] = sheet_no_str
	yinbi['AG3'] = sheet_no_str
	baoshen['K21'] = date_to_chinese(date)
	jianyan['AY26'] = date_to_chinese(date)
	yinbi['Y5'] = date_to_chinese(date)
	page = 0
	offset = 0
	first_col = 0
	last_page = 0
	uids = []
	for rows in v.values():
		for row in rows:
			first_col += 1
			last_page = page
			shigong.cell(row=get_row(page, offset), column=1).value = first_col
			for i in range(1, 12):
				shigong.cell(row=get_row(page, offset), column=i+1).value = row[i].value
			shigong.cell(row=get_row(page, offset), column=3).value = row[2].value.replace('.', '/')
			uids.append(row[3].value)
			offset += 1
			if offset >= 20:
				offset -= 20
				page += 1
		if offset != 0:
			offset = 0
			page += 1
	sample_n = sample_num(first_col)
	jianyan['BB8'] = first_col
	for i in range(12, 23):
		jianyan['AD{}'.format(i)] = sample_n
		jianyan['AI{}'.format(i)] = sample_n
		jianyan['AL{}'.format(i)] = '检查{}处，合格{}处'.format(sample_n, sample_n)
	print_area_str = 'A1:L{}'.format(last_page * 24 + 24)
	shigong.print_area = print_area_str
	uid_str = uids_to_str(uids)
	baoshen['C8'] = uid_str
	yinbi['G7'] = uid_str
	return print_area_str

def process(data, folder, prefix, fix=False):
	os.makedirs(folder, exist_ok=True)
	d = bin_by_date_and_id(data)
	sheet_no = 0
	all_in_one_wb = load_workbook('a.xlsx')
	area = dict()
	for date, v in d.items():
		sheet_no += 1
		template_wb = load_workbook('a.xlsx')
		ws1 = all_in_one_wb.copy_worksheet(all_in_one_wb['报审表'])
		ws1.title = '报审表{}'.format(sheet_no)
		# ws1.print_area = 'A1:L34'
		# ws1.print_area = template_wb['报审表'].print_area[0]
		ws2 = all_in_one_wb.copy_worksheet(all_in_one_wb['施工记录'])
		ws2.title = '施工记录{}'.format(sheet_no)
		# ws2.print_area = template_wb['施工记录'].print_area[0]
		ws3 = all_in_one_wb.copy_worksheet(all_in_one_wb['检验批'])
		ws3.title = '检验批{}'.format(sheet_no)
		# ws3.print_area = template_wb['检验批'].print_area[0]
		ws4 = all_in_one_wb.copy_worksheet(all_in_one_wb['隐蔽工程'])
		ws4.title = '隐蔽工程{}'.format(sheet_no)
		# ws4.print_area = template_wb['隐蔽工程'].print_area[0]
		modify(date, v, sheet_no, template_wb['报审表'], template_wb['施工记录'], template_wb['检验批'], template_wb['隐蔽工程'])
		area[sheet_no] = modify(date, v, sheet_no, ws1, ws2, ws3, ws4)
		template_wb.save('{}/{}.xlsx'.format(folder, prefix + date[5:]))
	all_in_one_wb.remove(all_in_one_wb['报审表'])
	all_in_one_wb.remove(all_in_one_wb['施工记录'])
	all_in_one_wb.remove(all_in_one_wb['检验批'])
	all_in_one_wb.remove(all_in_one_wb['隐蔽工程'])
	all_in_one_wb.save('{}/all.xlsx'.format(folder))
	if fix:
		fix_wb = load_workbook('{}/all.xlsx'.format(folder))
		for name in fix_wb.sheetnames:
			if name.startswith('报审表'):
				fix_wb[name].print_area = 'A1:L34'
			elif name.startswith('施工记录'):
				if len(name) != 4:
					fix_wb[name].print_area = area[int(name[4:])]
			elif name.startswith('检验批'):
				fix_wb[name].print_area = 'A1:BM31'
			else:
				fix_wb[name].print_area = 'A1:AL32'
		fix_wb.save('{}/all.xlsx'.format(folder))

args = parser.parse_args()
if not args.separate:
	data = reduce(lambda x, y: x + y, [list(data_wb[name].rows)[4:] for name in ids])
	data = filter(lambda x: x[0].value is not None, data)
	process(data, 'result', '')
else:
	for name in ids:
		data = list(data_wb[name].rows)[4:]
		data = filter(lambda x: x[0].value is not None, data)
		process(data, 'result/data' + name, name + '单轴检验批')

